#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
aggregate_signals_llm.py
(исправленная версия: константы не обрабатываются, только справочно)
"""

from __future__ import annotations

import argparse
import csv
import glob
import os
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------- #
# Константы / настройки по умолчанию
# --------------------------------------------------------------------------- #

DEFAULT_PROCESS_SHEETS = ["Правила и параметры"]
DEFAULT_REFERENCE_SHEETS = ["Константы"]

COL_KKS = "KKS код"
COL_DESC = "Описание"
COL_UNIT = "Ед. изм."
COL_USED = "Используемые сигналы"
COL_CODE = "Код"
COL_CALC_DESC = "Описание расчета"

REQUIRED_COLUMNS = [COL_KKS, COL_DESC, COL_UNIT, COL_USED, COL_CODE]

SYNTAX_2_MARKERS = [
    "enthalpy_ps",
    "enthalpy_pt",
    "pressure_saturation",
    "temperature_saturation",
    "entropy_pt",
    "temperature_ps",
]

CONST_SUFFIX = "CONST"
TOKEN_SPLIT_RE = re.compile(r"[;,\n|]+")

# --------------------------------------------------------------------------- #
# Вспомогательные структуры данных
# --------------------------------------------------------------------------- #


def estimate_complexity(code: str) -> str:
    """
    Возвращает один из уровней сложности: 'simple', 'medium', 'complex'.
    Оценка основана на количественных метриках кода.
    """
    if not code or not code.strip():
        return "simple"

    # Убираем пробелы и переводы строк для анализа
    clean = code.strip()
    length = len(clean)

    # Количество арифметических операторов (включая **)
    arithmetic_ops = len(re.findall(r'[+\-*/]|\*\*', clean))

    # Количество вызовов функций (идентификатор + '(')
    function_calls = len(re.findall(r'\b[A-Za-z_][A-Za-z0-9_]*\s*\(', clean))

    # Количество конструкций when (условных выражений)
    when_count = len(re.findall(r'\bwhen\b', clean, flags=re.IGNORECASE))

    # Количество HISTORY-операций (например, HISTORY_AVG, HISTORY_SUM и т.п.)
    history_count = len(re.findall(r'\bHISTORY[A-Za-z0-9_]*', clean, flags=re.IGNORECASE))

    # Максимальная вложенность скобок
    max_depth = 0
    current_depth = 0
    for ch in clean:
        if ch == '(':
            current_depth += 1
            max_depth = max(max_depth, current_depth)
        elif ch == ')':
            current_depth = max(0, current_depth - 1)

    # --- Пороговые значения (можно настраивать под свои нужды) ---
    # Для уровня "simple":
    SIMPLE_MAX_LENGTH = 100
    SIMPLE_MAX_ARITH = 4
    SIMPLE_MAX_FUNCS = 3
    SIMPLE_MAX_WHEN = 2
    SIMPLE_MAX_HISTORY = 1
    SIMPLE_MAX_DEPTH = 2

    # Для уровня "medium":
    MEDIUM_MAX_LENGTH = 400
    MEDIUM_MAX_ARITH = 8
    MEDIUM_MAX_FUNCS = 4
    MEDIUM_MAX_WHEN = 3
    MEDIUM_MAX_HISTORY = 2
    MEDIUM_MAX_DEPTH = 4

    # Вычисляем степень сложности на основе порогов
    if (length <= SIMPLE_MAX_LENGTH and
        arithmetic_ops <= SIMPLE_MAX_ARITH and
        function_calls <= SIMPLE_MAX_FUNCS and
        when_count <= SIMPLE_MAX_WHEN and
        history_count <= SIMPLE_MAX_HISTORY and
        max_depth <= SIMPLE_MAX_DEPTH):
        return "simple"
    elif (length <= MEDIUM_MAX_LENGTH and
          arithmetic_ops <= MEDIUM_MAX_ARITH and
          function_calls <= MEDIUM_MAX_FUNCS and
          when_count <= MEDIUM_MAX_WHEN and
          history_count <= MEDIUM_MAX_HISTORY and
          max_depth <= MEDIUM_MAX_DEPTH):
        return "medium"
    else:
        return "complex"

@dataclass
class SignalNode:
    kks: str
    sheet: str
    row: int
    description: str
    unit: str
    used_raw: str
    code: str
    deps: List[str] = field(default_factory=list)
    raw_signals: List[str] = field(default_factory=list)
    unresolved: List[str] = field(default_factory=list)
    used_by: List[str] = field(default_factory=list)
    syntax_version: int = 1
    calc_description: Optional[str] = None
    calc_col: int = 0
    is_reference: bool = False          # True для справочных листов (например, "Константы")

# --------------------------------------------------------------------------- #
# Загрузка текстовых файлов
# --------------------------------------------------------------------------- #

def read_text_file(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return path.read_text(encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue
    raise RuntimeError(f"Не удалось прочитать файл {path} ни в одной из кодировок")


def load_raw_signals(signals_dir: Path) -> Dict[str, str]:
    result: Dict[str, str] = {}
    csv_files = sorted(glob.glob(str(signals_dir / "*.csv")))
    if not csv_files:
        print(f"[WARN] В папке {signals_dir} не найдено csv-файлов с сигналами", file=sys.stderr)

    for csv_path in csv_files:
        text = read_text_file(Path(csv_path))
        reader = csv.DictReader(text.splitlines(), delimiter=";")
        if reader.fieldnames is None:
            continue
        norm_fields = {f.strip().lower(): f for f in reader.fieldnames}
        tag_col = norm_fields.get("tagname")
        desc_col = norm_fields.get("description")
        if not tag_col or not desc_col:
            print(f"[WARN] В файле {csv_path} не найдены столбцы Tagname/Description, пропускаю", file=sys.stderr)
            continue
        for row in reader:
            tag = (row.get(tag_col) or "").strip()
            desc = (row.get(desc_col) or "").strip()
            if not tag:
                continue
            if tag in result and result[tag] != desc:
                print(f"[WARN] Сигнал {tag} повторяется с другим описанием в {csv_path} — оставляю первое найденное", file=sys.stderr)
                continue
            result[tag] = desc
    print(f"[INFO] Загружено {len(result)} сигналов АСУ ТП из {len(csv_files)} csv-файлов")
    return result

# --------------------------------------------------------------------------- #
# Разбор xlsx
# --------------------------------------------------------------------------- #

def normalize_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def find_header_columns(ws: Worksheet, header_row: int = 1) -> Dict[str, int]:
    expected_norm = {normalize_header(c): c for c in REQUIRED_COLUMNS + [COL_CALC_DESC]}
    found: Dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        norm = normalize_header(str(cell.value))
        if norm in expected_norm:
            found[expected_norm[norm]] = cell.column

    missing = [c for c in REQUIRED_COLUMNS if c not in found]
    if missing:
        raise RuntimeError(
            f"На листе '{ws.title}' не найдены обязательные столбцы: {missing}. "
            f"Найдены заголовки: {[c.value for c in ws[header_row]]}"
        )
    return found


def ensure_calc_desc_column(ws: Worksheet, columns: Dict[str, int], header_row: int = 1) -> int:
    if COL_CALC_DESC in columns:
        return columns[COL_CALC_DESC]
    new_col = ws.max_column + 1
    header_cell = ws.cell(row=header_row, column=new_col, value=COL_CALC_DESC)
    header_cell.font = Font(bold=True)
    columns[COL_CALC_DESC] = new_col
    return new_col


def parse_sheet(ws: Worksheet, header_row: int = 1, mark_as_reference: bool = False) -> List[SignalNode]:
    columns = find_header_columns(ws, header_row)
    calc_col = ensure_calc_desc_column(ws, columns, header_row)

    nodes: List[SignalNode] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        kks_val = ws.cell(row=row_idx, column=columns[COL_KKS]).value
        kks = (str(kks_val).strip() if kks_val is not None else "")
        if not kks:
            continue

        desc = ws.cell(row=row_idx, column=columns[COL_DESC]).value or ""
        unit = ws.cell(row=row_idx, column=columns[COL_UNIT]).value or ""
        used = ws.cell(row=row_idx, column=columns[COL_USED]).value or ""
        code = ws.cell(row=row_idx, column=columns[COL_CODE]).value or ""
        existing_calc = ws.cell(row=row_idx, column=calc_col).value

        node = SignalNode(
            kks=kks,
            sheet=ws.title,
            row=row_idx,
            description=str(desc).strip(),
            unit=str(unit).strip(),
            used_raw=str(used).strip(),
            code=str(code).strip(),
            calc_col=calc_col,
            calc_description=(str(existing_calc).strip() if existing_calc else None),
            is_reference=mark_as_reference,
        )
        nodes.append(node)
    return nodes

# --------------------------------------------------------------------------- #
# Построение графа зависимостей
# --------------------------------------------------------------------------- #

def split_used_signals(text: str) -> List[str]:
    if not text:
        return []
    tokens = [t.strip() for t in TOKEN_SPLIT_RE.split(text)]
    return [t for t in tokens if t]


def build_graph(nodes: List[SignalNode], raw_signals: Dict[str, str]) -> Dict[str, SignalNode]:
    registry: Dict[str, SignalNode] = {n.kks: n for n in nodes}

    for node in nodes:
        for token in split_used_signals(node.used_raw):
            if token == node.kks:
                continue
            if token in registry:
                node.deps.append(token)
            elif token in raw_signals:
                node.raw_signals.append(token)
            else:
                node.unresolved.append(token)
                suffix_hint = " (похоже на константу, но не найдена на листе 'Константы')" if token.upper().endswith(CONST_SUFFIX) else ""
                print(f"[WARN] Сигнал {node.kks}: не найден используемый сигнал '{token}'{suffix_hint}", file=sys.stderr)
    for node in nodes:
        for dep_kks in node.deps:
            if dep_kks in registry:
                registry[dep_kks].used_by.append(node.kks)

    return registry


def topological_levels(nodes: List[SignalNode]) -> List[List[str]]:
    """
    Принимает список узлов, для которых нужно построить порядок.
    Учитываются только зависимости между этими узлами.
    """
    registry = {n.kks: n for n in nodes}
    remaining = {}
    for kks, node in registry.items():
        filtered_deps = set(d for d in node.deps if d in registry)
        remaining[kks] = filtered_deps

    levels: List[List[str]] = []
    while remaining:
        current_level = [kks for kks, deps in remaining.items() if not deps]
        if not current_level:
            raise RuntimeError(
                "Обнаружен цикл в зависимостях синтетических сигналов, "
                f"не удалось разложить по уровням: {list(remaining.keys())}"
            )
        levels.append(sorted(current_level))
        for kks in current_level:
            del remaining[kks]
        for deps in remaining.values():
            deps.difference_update(current_level)

    return levels

# --------------------------------------------------------------------------- #
# Определение версии синтаксиса
# --------------------------------------------------------------------------- #

def detect_syntax_version(code: str) -> int:
    for marker in SYNTAX_2_MARKERS:
        if re.search(rf"\b{re.escape(marker)}\b", code):
            return 2
    return 1

# --------------------------------------------------------------------------- #
# Построение "матрёшки"
# --------------------------------------------------------------------------- #

def render_hierarchy(node: SignalNode, registry: Dict[str, SignalNode], depth: int = 0, seen: Optional[set] = None) -> str:
    if seen is None:
        seen = set()
    lines: List[str] = []
    indent = "  " * depth

    for dep_kks in node.deps:
        dep = registry.get(dep_kks)
        if dep is None:
            continue
        if dep.sheet == "Константы":
            continue  # константы выводятся отдельным блоком
        marker = f"{indent}- [{dep.kks}] {dep.description} (Ед. изм.: {dep.unit or '—'})"
        lines.append(marker)
        if dep.calc_description:
            lines.append(f"{indent}  Описание расчёта: {dep.calc_description}")
        if dep.code:
            lines.append(f"{indent}  Код:")
            for code_line in dep.code.splitlines():
                lines.append(f"{indent}    {code_line}")
        if dep_kks not in seen:
            seen.add(dep_kks)
            nested = render_hierarchy(dep, registry, depth + 1, seen)
            if nested:
                lines.append(nested)

    return "\n".join(lines)

# --------------------------------------------------------------------------- #
# Формирование промпта и вызов Ollama
# --------------------------------------------------------------------------- #

SYSTEM_PROMPT_TEMPLATE = """\
Ты — опытный инженер-теплотехник и специалист по программированию расчётных \
синтетических сигналов для АСУ ТП энергетического оборудования (турбины, ПГУ, \
котлы, конденсационные установки). Тебе на вход даётся описание одного \
синтетического сигнала: его код, входные сигналы и, при наличии, уже готовые \
описания вложенных (используемых) синтетических сигналов.

Твоя задача — написать для столбца "Описание расчета" описание принципа расчёта \
ИМЕННО этого сигнала. В конце промпта будет указана требуемая степень детализации; \
строго следуй ей. В описании обязательно учитывай (в зависимости от сложности):
  1. Физический смысл сигнала — что он показывает и зачем нужен.
  2. Пошаговое объяснение алгоритма расчёта по коду: что вычисляется на каждом \
значимом шаге, в каком порядке, с какими входами.
  3. При необходимости — математическое и физическое описание используемых \
формул/термодинамических функций (например, если используются функции \
энтальпии, энтропии, параметров насыщения и т.п. — поясни физический смысл \
этих величин и то, как они используются здесь).
  4. Если сигнал использует другие (вложенные) синтетические сигналы — не \
нужно повторно объяснять, КАК они сами рассчитываются (это уже описано \
отдельно и дано тебе для контекста), но обязательно объясни, ЧТО каждый из \
них означает и какую роль он играет в расчёте текущего сигнала.
  5. Если используются константы (сигналы, оканчивающиеся на CONST) — поясни, \
что это за константа и как она участвует в расчёте.

Пиши на русском языке, техническим, но понятным языком, связным текстом (не \
просто списком). Строго придерживайся синтаксиса кода, описанного в \
приложенных материалах, при интерпретации операторов и функций — не \
придумывай функции или семантику, которых нет в этих материалах.

ВАЖНО: Избегай субъективных оценок, рекламных эпитетов и усилительных \
прилагательных, таких как «жёсткий», «критически важный», «уникальный», \
«высокоточный», «надёжный», «эффективный» и им подобных. Используй только \
нейтральные технические формулировки. Если нужно подчеркнуть важность \
какого-либо условия, опиши его фактическую роль в алгоритме без эмоциональной \
окраски.

СПИСОК некоторых сокращений: \
ПСГ - подогреватель сетевой воды\
ЧВД - часть высокого давления (у паровой турбины)\
ЧСД - часть среднего давления (у паровой турбины)\
ЦВД - цилиндр выского давления\
ЦСД - цилиндр среднего давления\
ЦНД - цилиндр низкого давления
"""


def build_prompt(node: SignalNode, registry: Dict[str, SignalNode], raw_signals: Dict[str, str],
                  syntax_texts: Dict[int, str], kks_text: str) -> Tuple[str, str]:
    parts: List[str] = []

    parts.append("## Синтаксис кода (обязателен к соблюдению)\n")
    parts.append(syntax_texts[node.syntax_version])
    parts.append("\n\n## Принципы формирования KKS-сигналов\n")
    parts.append(kks_text)

    if node.raw_signals:
        parts.append("\n\n## Используемые сигналы АСУ ТП (входные, измеряемые)\n")
        for tag in node.raw_signals:
            parts.append(f"- {tag}: {raw_signals.get(tag, '(описание не найдено)')}\n")

    const_only = [d for d in node.deps if registry.get(d) and registry[d].sheet == "Константы"]
    if const_only:
        parts.append("\n\n## Используемые константы (лист 'Константы')\n")
        for c_kks in const_only:
            c = registry[c_kks]
            parts.append(f"- [{c.kks}] {c.description} (Ед. изм.: {c.unit or '—'})"
                          f"{'; ' + c.calc_description if c.calc_description else ''}\n")

    non_const_deps = [d for d in node.deps if registry.get(d) and registry[d].sheet != "Константы"]
    if non_const_deps:
        hierarchy_text = render_hierarchy(node, registry)
        if hierarchy_text:
            parts.append("\n\n## Иерархия вложенных синтетических сигналов, используемых в расчёте\n")
            parts.append(hierarchy_text)

    if node.unresolved:
        parts.append("\n\n## Внимание: не найдены описания следующих сигналов (учитывай их только по названию)\n")
        for t in node.unresolved:
            parts.append(f"- {t}\n")

    if node.used_by:
        parents = [registry[p] for p in node.used_by if p in registry]
        if parents:
            parts.append("\n\n## Где используется этот сигнал (непосредственные потребители). Эта информация дана только для контекста и не должна пересказываться, но может быть использована для уточнения физического смысла \n")
            # Ограничим число, чтобы не перегружать
            max_parents = 10
            for parent in parents[:max_parents]:
                parts.append(f"- {parent.kks}: {parent.description}")
                # если у родителя уже есть сгенерированное описание расчёта, можно добавить кратко, но не обязательно
                # if parent.calc_description:
                #     parts.append(f"  Назначение: {parent.calc_description[:200]}...")
                parts.append("\n")

    parts.append("\n\n## Сигнал, для которого нужно составить 'Описание расчета'\n")
    parts.append(f"KKS код: {node.kks}\n")
    parts.append(f"Описание: {node.description}\n")
    parts.append(f"Ед. изм.: {node.unit or '—'}\n")
    parts.append("Код расчёта:\n")
    parts.append(node.code or "(код не задан)")

        # Оценка сложности кода
    complexity = estimate_complexity(node.code)

    # Формируем финальную инструкцию в зависимости от сложности
    if complexity == "simple":
        parts.append(
            f"\n\nКод этого сигнала {node.kks} очень простой (простые арифметические операции или "
            "одиночная функция). Напиши КРАТКОЕ описание расчёта: 2–3 предложения, "
            "только суть — что вычисляется и по какой формуле. Без лишних физических "
            "рассуждений и теоретических отступлений."
        )
    elif complexity == "medium":
        parts.append(
            f"\n\nКод сигнала {node.kks} средней сложности. Напиши описание расчёта среднего объёма: "
            "объясни алгоритм, укажи физический смысл, но не углубляйся в излишнюю "
            "детализацию. Достаточно 4–6 предложений."
        )
    else:  # complex
        parts.append(
            f"\n\nКод сигнала {node.kks} сложный. Напиши ПОДРОБНОЕ описание расчёта: пошаговый алгоритм, "
            "физическое и математическое обоснование, объяснение всех используемых "
            "функций и переменных. Пиши развёрнуто, связным текстом."
        )

    parts.append(
        "\n\nДай только сам текст описания, без заголовков вроде 'Описание расчета:' "
        "и без markdown-разметки."
    )

    return SYSTEM_PROMPT_TEMPLATE, "".join(parts)


def call_ollama(base_url: str, model: str, system: str, prompt: str,
                 temperature: float, timeout: int, max_retries: int = 3) -> str:
    url = base_url.rstrip("/") + "/api/generate"
    payload = {
        "model": model,
        "system": system,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": temperature},
    }

    last_error: Optional[Exception] = None
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(url, json=payload, timeout=timeout)
            resp.raise_for_status()
            data = resp.json()
            return (data.get("response") or "").strip()
        except Exception as exc:
            last_error = exc
            print(f"[WARN] Попытка {attempt}/{max_retries} вызова Ollama для не удалась: {exc}", file=sys.stderr)
            time.sleep(min(5 * attempt, 20))
    raise RuntimeError(f"Не удалось получить ответ от Ollama после {max_retries} попыток: {last_error}")


# --------------------------------------------------------------------------- #
# Прогресс
# --------------------------------------------------------------------------- #

def print_progress(level_idx: int, levels_total: int, item_idx: int, level_size: int,
                    total_done: int, total_all: int, kks: str, status: str, elapsed: float = 0.0) -> None:
    msg = (f"[Уровень {level_idx + 1}/{levels_total}] [{item_idx}/{level_size}] "
           f"[Всего {total_done}/{total_all}] {kks}: {status}")
    if elapsed:
        msg += f" ({elapsed:.1f} с)"
    print(msg, flush=True)


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", type=Path, help="Путь к xlsx-таблице синтетических сигналов", default='Params.xlsx')
    ap.add_argument("--syntax-dir", type=Path, help="Папка с syntax_1.md и syntax_2.md", default='syntax')
    ap.add_argument("--syntax1-name", default="syntax_1.md")
    ap.add_argument("--syntax2-name", default="syntax_2.md")
    ap.add_argument("--kks", type=Path, help="Путь к kks.md", default='kks.md')
    ap.add_argument("--signals-dir", type=Path, help="Папка signals с csv-файлами сигналов АСУ ТП", default='signals_folder')
    ap.add_argument("--sheets", nargs="+", default=DEFAULT_PROCESS_SHEETS,
                    help="Листы для обработки (по умолчанию только 'Правила и параметры')")
    ap.add_argument("--reference-sheets", nargs="+", default=DEFAULT_REFERENCE_SHEETS,
                    help="Справочные листы, которые не обрабатываются, но используются для разрешения зависимостей")
    ap.add_argument("--header-row", type=int, default=1)
    ap.add_argument("--model", default="gemma4:31b")
    ap.add_argument("--ollama-url", default="http://192.168.0.5:11434")
    ap.add_argument("--temperature", type=float, default=0.4)
    ap.add_argument("--timeout", type=int, default=600, help="Таймаут одного вызова Ollama, сек")
    ap.add_argument("--overwrite", action="store_true", help="Пересчитать все описания заново")
    ap.add_argument("--limit", type=int, default=None, help="Обработать не более N сигналов (для теста)")
    ap.add_argument("--save-every", type=int, default=1, help="Сохранять xlsx каждые N обработанных сигналов")
    ap.add_argument("--dry-run", action="store_true", help="Не вызывать LLM, только показать план обработки")
    ap.add_argument("--output", type=Path, default=None, help="Путь для сохранения результата (по умолчанию — перезаписать --xlsx)")
    args = ap.parse_args()

    output_path = args.output or args.xlsx

    # --- Загрузка вспомогательных материалов ------------------------------ #
    syntax_texts = {
        1: read_text_file(args.syntax_dir / args.syntax1_name),
        2: read_text_file(args.syntax_dir / args.syntax2_name),
    }
    kks_text = read_text_file(args.kks)
    raw_signals = load_raw_signals(args.signals_dir)

    # --- Загрузка книги и построение реестра узлов ------------------------- #
    wb = load_workbook(args.xlsx)
    all_nodes: List[SignalNode] = []

    # 1. Загружаем обрабатываемые листы
    for sheet_name in args.sheets:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Лист '{sheet_name}' не найден в книге. Доступные листы: {wb.sheetnames}")
        ws = wb[sheet_name]
        sheet_nodes = parse_sheet(ws, args.header_row)
        print(f"[INFO] Обрабатываемый лист '{sheet_name}': найдено {len(sheet_nodes)} сигналов")
        all_nodes.extend(sheet_nodes)

    # 2. Загружаем справочные листы (обычно "Константы")
    for ref_sheet in args.reference_sheets:
        if ref_sheet in wb.sheetnames and ref_sheet not in args.sheets:
            ws = wb[ref_sheet]
            ref_nodes = parse_sheet(ws, args.header_row, mark_as_reference=True)
            print(f"[INFO] Справочный лист '{ref_sheet}': загружено {len(ref_nodes)} записей (обработка не требуется)")
            all_nodes.extend(ref_nodes)

    # Построение общего реестра
    registry = build_graph(all_nodes, raw_signals)

    # Вычисляем версии синтаксиса для всех узлов
    for node in all_nodes:
        node.syntax_version = detect_syntax_version(node.code)

    # Для обработки берём только узлы обрабатываемых листов
    processing_nodes = [n for n in all_nodes if n.sheet in args.sheets]
    if not processing_nodes:
        raise RuntimeError("Нет ни одного сигнала для обработки. Проверьте --sheets.")

    # Топологическая сортировка только по обрабатываемым узлам
    levels = topological_levels(processing_nodes)
    total_all = sum(len(l) for l in levels)
    print(f"[INFO] Построена иерархия из {len(levels)} уровней, всего {total_all} сигналов для обработки")
    for i, level in enumerate(levels):
        print(f"        Уровень {i + 1}: {len(level)} сигнал(ов) -> {level}")

    # --- Обработка по уровням ---------------------------------------------- #
    total_done = 0
    processed_since_save = 0
    start_time = time.time()

    for level_idx, level in enumerate(levels):
        for item_idx, kks in enumerate(level, start=1):
            node = registry[kks]  # полный узел (все данные на месте)
            total_done += 1

            if args.limit is not None and total_done > args.limit:
                print("[INFO] Достигнут --limit, останавливаюсь")
                break

            if node.calc_description and not args.overwrite:
                print_progress(level_idx, len(levels), item_idx, len(level),
                                total_done, total_all, kks, "уже заполнено, пропускаю")
                continue

            system_prompt, user_prompt = build_prompt(node, registry, raw_signals, syntax_texts, kks_text)

            if args.dry_run:
                print_progress(level_idx, len(levels), item_idx, len(level),
                                total_done, total_all, kks,
                                f"[dry-run] синтаксис v{node.syntax_version}, "
                                f"длина промпта: {len(user_prompt)} симв., "
                                f"зависимостей: {len(node.deps)}, сырых сигналов: {len(node.raw_signals)}")
                continue

            t0 = time.time()
            print_progress(level_idx, len(levels), item_idx, len(level),
                            total_done, total_all, kks, "отправка запроса в Ollama...")
            try:
                result = call_ollama(args.ollama_url, args.model, system_prompt, user_prompt,
                                      args.temperature, args.timeout)
            except Exception as exc:
                print_progress(level_idx, len(levels), item_idx, len(level),
                                total_done, total_all, kks, f"ОШИБКА: {exc}", time.time() - t0)
                continue

            node.calc_description = result
            elapsed = time.time() - t0
            print_progress(level_idx, len(levels), item_idx, len(level),
                            total_done, total_all, kks,
                            f"готово, {len(result)} симв.", elapsed)

            ws = wb[node.sheet]
            ws.cell(row=node.row, column=node.calc_col, value=node.calc_description)

            processed_since_save += 1
            if processed_since_save >= args.save_every:
                wb.save(output_path)
                processed_since_save = 0

        else:
            continue
        break  # прерываем внешний цикл, если сработал --limit

    if not args.dry_run:
        wb.save(output_path)
        total_elapsed = time.time() - start_time
        print(f"[INFO] Готово. Обработано сигналов: {total_done}/{total_all}. "
              f"Затрачено времени: {total_elapsed:.1f} с. Файл сохранён: {output_path}")
    else:
        print("[INFO] Dry-run завершён, файл не изменялся.")


if __name__ == "__main__":
    main()