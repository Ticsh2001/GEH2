# main.py — версия с поддержкой конфигураций

import os
import re
import json
import uuid
import pickle
import tempfile
from typing import Dict, List, Any, Optional
from io import BytesIO
from update_projects import update_projects_if_templates_changed
from datetime import datetime
import io
import requests
import httpx

import nn.training_queue as tq
import nn.nn_template as nn_template

from dataprocessing import process_element, get_datasets_dir, get_dataset_path, get_meta_path, get_output_file, get_element_status




from openpyxl import Workbook
from openpyxl.styles import Alignment

import pandas as pd
from fastapi import FastAPI, HTTPException, Request, Query, File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import shutil
from fastapi import Body
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from urllib.parse import quote
from export import export_selected_projects
from export import get_code_length

# =============================================================================
# КОНФИГУРАЦИЯ
# =============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_PATH = os.path.join(BASE_DIR, "settings.json")
TEMPLATES_PATH = os.path.join(BASE_DIR, "formula_templates.json")
NN_BLOCK_PARAMS_PATH = os.path.join(BASE_DIR, "nn_block_params.json")

SIGNAL_INDEX_CACHE = {}          # кэш индексов сигналов: config -> (folder_state, index)
SIGNALS_CACHE = {}              # кэш списков сигналов: config -> list
TABLES_CACHE = {}               # кэш списка таблиц: config -> list

# =============================================================================
# PYDANTIC МОДЕЛИ
# =============================================================================

class VisualizerStateRequest(BaseModel):
    session_token: str
    state: Dict[str, Any]

class VisualizerStateResponse(BaseModel):
    success: bool
    state: Optional[Dict[str, Any]] = None
    message: Optional[str] = None

class VisualizeSessionRequest(BaseModel):
    signals: List[str]
    code: str = ""
    visualizer_state: Optional[Dict[str, Any]] = None

# =============================================================================
# ГЛОБАЛЬНОЕ СОСТОЯНИЕ
# =============================================================================

STATE = {
    "settings": None,
    "signals": None,          # больше не глобальный, оставлен для обратной совместимости (не используется)
    "templates": None,
    "configurations": [],     # список имён конфигураций
}

TAGS_FILE = "tags.json"

def load_tags_data():
    if not os.path.exists(TAGS_FILE):
        return {"tags": [], "assignments": {}}
    try:
        with open(TAGS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"tags": [], "assignments": {}}

def save_tags_data(data):
    with open(TAGS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# -----------------------------------------------------------------------------
# Вспомогательные функции для работы с путями конфигураций
# -----------------------------------------------------------------------------

def _abs_folder(setting_key: str) -> Optional[str]:
    """Абсолютный путь к базовой папке из настроек (без учёта конфигурации)."""
    base = STATE["settings"].get(setting_key)
    if not base:
        return None
    if not os.path.isabs(base):
        base = os.path.normpath(os.path.join(BASE_DIR, base))
    return base

def config_path(setting_key: str, config: str) -> Optional[str]:
    """Возвращает абсолютный путь к подпапке <config> внутри базовой папки."""
    base = _abs_folder(setting_key)
    if not base:
        return None
    return os.path.join(base, config) if config else base

def ensure_config_dirs(config: str):
    """Создаёт подпапки конфигурации во всех рабочих папках, если их ещё нет."""
    for key in ("projectDataFolder", "signalDataFolder", "signalArchiveFolder",
                "tablesFolder", "deletedFolder"):  # добавим deletedFolder
        base = _abs_folder(key)
        if base:
            full = os.path.join(base, config)
            os.makedirs(full, exist_ok=True)

# -----------------------------------------------------------------------------
# Работа с таблицами
# -----------------------------------------------------------------------------

def load_tables_from_folder(folder: str) -> List[Dict]:
    folder_abs = folder if os.path.isabs(folder) else os.path.normpath(os.path.join(BASE_DIR, folder))
    if not os.path.isdir(folder_abs):
        return []
    items = []
    for name in os.listdir(folder_abs):
        if not name.lower().endswith(".xlsx"):
            continue
        base_name = os.path.splitext(name)[0]
        items.append({"Name": base_name, "Description": ""})
    items.sort(key=lambda x: x["Name"].lower())
    return items

def load_tables_meta(folder: str) -> Dict[str, str]:
    folder_abs = folder if os.path.isabs(folder) else os.path.normpath(os.path.join(BASE_DIR, folder))
    meta_path = os.path.join(folder_abs, "tables.json")
    if not os.path.isfile(meta_path):
        return {}
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        meta = {}
        if isinstance(data, list):
            for obj in data:
                if isinstance(obj, dict) and obj:
                    k, v = next(iter(obj.items()))
                    meta[str(k)] = str(v)
        elif isinstance(data, dict):
            meta = {str(k): str(v) for k, v in data.items()}
        return meta
    except Exception as e:
        print(f"[WARN] failed to read tables.json: {e}")
        return {}

def get_tables_for_config(config: str) -> List[Dict]:
    """Возвращает список таблиц для заданной конфигурации, используя кэш."""
    if config in TABLES_CACHE:
        return TABLES_CACHE[config]
    folder = config_path("tablesFolder", config)
    if not folder:
        return []
    items = load_tables_from_folder(folder)
    meta = load_tables_meta(folder)
    for item in items:
        name = item["Name"]
        if name in meta:
            item["Description"] = meta[name]
    TABLES_CACHE[config] = items
    return items

# -----------------------------------------------------------------------------
# Загрузка настроек и шаблонов
# -----------------------------------------------------------------------------

def load_settings() -> Dict:
    with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def load_templates() -> Dict:
    if not os.path.exists(TEMPLATES_PATH):
        return {"templates": []}
    with open(TEMPLATES_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

# -----------------------------------------------------------------------------
# Сигналы (базовые + проектные)
# -----------------------------------------------------------------------------

def load_signals_from_folder(folder: str) -> List[Dict]:
    folder_abs = folder if os.path.isabs(folder) else os.path.normpath(os.path.join(BASE_DIR, folder))
    if not os.path.isdir(folder_abs):
        return []
    signals_map = {}
    for name in os.listdir(folder_abs):
        if not name.lower().endswith(".csv"):
            continue
        path = os.path.join(folder_abs, name)
        try:
            try:
                df = pd.read_csv(path, sep=';')[['Tagname', 'Description', 'Engineering Unit']]
            except KeyError:
                df = pd.read_csv(path, sep=';')[['Tagname', 'Description']]
            df = df.dropna(subset=['Tagname'])
            for _, row in df.iterrows():
                tag = str(row['Tagname']).strip()
                desc = "" if pd.isna(row['Description']) else str(row['Description']).strip()
                try:
                    unit = "" if pd.isna(row['Engineering Unit']) else str(row['Engineering Unit']).strip()
                except KeyError:
                    unit = ""
                desc_full = ", ".join([x for x in [desc, unit] if x])
                if tag:
                    signals_map[tag] = {
                        "Tagname": tag,
                        "Description": desc_full,
                        "EngineeringUnit": unit
                    }
        except Exception as e:
            print(f"[WARN] failed to read {path}: {e}")
    out = list(signals_map.values())
    out.sort(key=lambda x: x["Tagname"])
    return out

def load_project_signals(folder: str) -> List[Dict]:
    folder_abs = folder if os.path.isabs(folder) else os.path.normpath(os.path.join(BASE_DIR, folder))
    if not os.path.isdir(folder_abs):
        return []
    out = []
    for name in os.listdir(folder_abs):
        if not name.endswith(".json"):
            continue
        path = os.path.join(folder_abs, name)
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            proj = payload.get("project", {}) or {}
            code = (proj.get("code") or "").strip()
            if not code:
                continue
            desc = (proj.get("description") or "").strip()
            dim = (proj.get("dimension") or "").strip()
            out.append({
                "Tagname": code,
                "Description": desc,
                "EngineeringUnit": dim,
                "Type": proj.get("type", "")
            })
        except Exception as e:
            print(f"[WARN] failed to read project {path}: {e}")
            continue
    out.sort(key=lambda x: x["Tagname"])
    return out

def get_signals_for_config(config: str) -> List[Dict]:
    """Возвращает объединённый список сигналов для конфигурации, с кэшированием."""
    if config in SIGNALS_CACHE:
        return SIGNALS_CACHE[config]
    base_folder = config_path("signalDataFolder", config)
    proj_folder = config_path("projectDataFolder", config)
    base = load_signals_from_folder(base_folder) if base_folder else []
    proj = load_project_signals(proj_folder) if proj_folder else []
    merged = {}
    for s in base:
        merged[s["Tagname"]] = s
    for s in proj:
        merged[s["Tagname"]] = s
    out = list(merged.values())
    out.sort(key=lambda x: x["Tagname"])
    SIGNALS_CACHE[config] = out
    return out

def invalidate_signals_cache(config: str):
    """Сбрасывает кэш сигналов для конкретной конфигурации."""
    SIGNALS_CACHE.pop(config, None)

# -----------------------------------------------------------------------------
# Индекс сигналов (архив)
# -----------------------------------------------------------------------------

def build_signal_index(folder: str) -> Dict[str, List[str]]:
    folder_abs = folder if os.path.isabs(folder) else os.path.normpath(os.path.join(BASE_DIR, folder))
    if not os.path.isdir(folder_abs):
        return {}
    signal_index = {}
    print(f"[INFO] Building signal index from {folder_abs}...")
    for filename in os.listdir(folder_abs):
        if not filename.lower().endswith(".csv"):
            continue
        filepath = os.path.join(folder_abs, filename)
        try:
            df_header = pd.read_csv(filepath, nrows=0, encoding="ISO-8859-2", sep=";")
            columns = df_header.columns.tolist()
            signal_columns = [c for c in columns if c not in ["DATE", "TIME", "datetime"]]
            for signal_name in signal_columns:
                signal_index.setdefault(signal_name, []).append(filepath)
        except Exception as e:
            print(f"  ✗ Failed to index {filename}: {e}")
            continue
    return signal_index

def get_folder_state(path: str) -> dict:
    if not os.path.isdir(path):
        return {}
    state = {}
    for name in os.listdir(path):
        if name.lower().endswith(".csv"):
            filepath = os.path.join(path, name)
            state[name] = os.path.getmtime(filepath)
    return state

def load_signal_index(config: str) -> Dict[str, List[str]]:
    """Загружает индекс сигналов для конфигурации, используя кэш в памяти."""
    archive = config_path("signalArchiveFolder", config)
    if not archive:
        return {}
    current_state = get_folder_state(archive)
    cached_entry = SIGNAL_INDEX_CACHE.get(config)
    if cached_entry:
        cached_state, cached_index = cached_entry
        if cached_state == current_state:
            return cached_index
    index = build_signal_index(archive)
    SIGNAL_INDEX_CACHE[config] = (current_state, index)
    return index

def load_signal_data_optimized(signal_names: List[str], config: str) -> Dict[str, pd.DataFrame]:
    archive = config_path("signalArchiveFolder", config)
    if not archive:
        raise RuntimeError("signalArchiveFolder not configured")
    signal_index = load_signal_index(config)
    signal_names_set = set(signal_names)
    files_to_load = set()
    for signal_name in signal_names_set:
        if signal_name in signal_index:
            files_to_load.update(signal_index[signal_name])
    found_signals = {}
    for filepath in files_to_load:
        try:
            df = pd.read_csv(filepath, encoding="ISO-8859-2", sep=";")
            df["TIME"] = df["TIME"].str.replace(",", ".", regex=False)
            df["TIME"] = df["TIME"].str.split(".").str[0]
            combined = df["DATE"] + " " + df["TIME"]
            df["datetime"] = pd.to_datetime(combined, format="%d.%m.%Y %H:%M:%S", errors="coerce")
            df = df.dropna(subset=["datetime"])
            df = df.drop(['DATE', 'TIME'], axis=1)
            df = df.sort_values("datetime")
            available_columns = set(df.columns) & signal_names_set
            for signal_name in available_columns:
                if signal_name not in found_signals:
                    found_signals[signal_name] = df[["datetime", signal_name]].copy()
                    found_signals[signal_name].columns = ["datetime", "value"]
        except Exception as e:
            print(f"[WARN] Failed to read {filepath}: {e}")
            continue
    return found_signals

# -----------------------------------------------------------------------------
# Проекты и зависимости
# -----------------------------------------------------------------------------

def get_project_path(filename: str, config: str) -> str:
    folder = config_path("projectDataFolder", config)
    if not folder:
        raise RuntimeError("projectDataFolder not configured")
    if '..' in filename or '/' in filename or '\\' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    path = os.path.join(folder, filename)
    if not path.startswith(folder):
        raise HTTPException(status_code=400, detail="Path traversal attempt")
    return path

def get_storage_path(filename: str, storage: str = "projects", config: str = None) -> str:
    key_map = {
        "projects": "projectDataFolder",
        "templates": "templateDataFolder"
    }
    key = key_map.get(storage)
    if not key:
        raise HTTPException(status_code=400, detail="Unknown storage")
    if storage == "templates":
        # шаблоны общие, не зависят от конфигурации
        base_dir = _abs_folder(key)
    else:
        base_dir = config_path(key, config)
    if not base_dir:
        raise HTTPException(status_code=500, detail=f"{key} not configured")
    os.makedirs(base_dir, exist_ok=True)
    if '..' in filename or '/' in filename or '\\' in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    path = os.path.join(base_dir, filename)
    if not path.startswith(base_dir):
        raise HTTPException(status_code=400, detail="Path traversal attempt")
    return path

def extract_input_signals_from_project(project_data: Dict) -> List[str]:
    elements = project_data.get("elements", {})
    input_signals = []
    for elem_id, elem_data in elements.items():
        if elem_data.get("type") == "input-signal":
            props = elem_data.get("props", {})
            signal_name = props.get("name")
            if signal_name:
                input_signals.append(signal_name)
    return input_signals

def load_project_by_code(code: str, config: str) -> Optional[Dict]:
    folder = config_path("projectDataFolder", config)
    if not folder or not os.path.isdir(folder):
        return None
    for name in os.listdir(folder):
        if not name.endswith(".json"):
            continue
        path = os.path.join(folder, name)
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            proj = payload.get("project", {})
            if proj.get("code") == code:
                return {
                    "project": proj,
                    "formula": payload.get("code", ""),
                    "elements": payload.get("elements", {})
                }
        except Exception as e:
            print(f"[WARN] Error reading project {path}: {e}")
            continue
    return None

def is_base_signal(signal_name: str, config: str) -> bool:
    signal_index = load_signal_index(config)
    return signal_name in signal_index

def resolve_signal_dependencies(
    signal_names: List[str],
    config: str,
    visited: set = None,
    resolved: Dict[str, Dict] = None
) -> tuple[set, Dict[str, Dict]]:
    if visited is None:
        visited = set()
    if resolved is None:
        resolved = {}
    base_signals = set()
    for signal_name in signal_names:
        if not signal_name or signal_name in visited:
            continue
        visited.add(signal_name)
        if is_base_signal(signal_name, config):
            base_signals.add(signal_name)
            continue
        project = load_project_by_code(signal_name, config)
        if project is None:
            base_signals.add(signal_name)
            continue
        formula = project.get("formula", "")
        dependencies = extract_input_signals_from_project(project)
        resolved[signal_name] = {
            "formula": formula,
            "dependencies": dependencies
        }
        sub_base, _ = resolve_signal_dependencies(dependencies, config, visited, resolved)
        base_signals.update(sub_base)
    return base_signals, resolved

def topological_sort_signals(synthetic_signals: Dict[str, Dict]) -> List[str]:
    if not synthetic_signals:
        return []
    in_degree = {name: 0 for name in synthetic_signals}
    graph = {name: [] for name in synthetic_signals}
    for name, data in synthetic_signals.items():
        for dep in data.get("dependencies", []):
            if dep == name:
                continue
            if dep in synthetic_signals:
                graph[dep].append(name)
                in_degree[name] += 1
    queue = [name for name, degree in in_degree.items() if degree == 0]
    result = []
    while queue:
        node = queue.pop(0)
        result.append(node)
        for neighbor in graph[node]:
            in_degree[neighbor] -= 1
            if in_degree[neighbor] == 0:
                queue.append(neighbor)
    if len(result) != len(synthetic_signals):
        cyclic = [name for name in synthetic_signals if name not in result]
        raise ValueError(f"Циклическая зависимость между сигналами: {cyclic}")
    return result

def upsert_formula_template_from_project(content: Dict[str, Any]) -> None:
    project_meta = content.get("project") or {}
    template_name = project_meta.get("code", "").strip()
    if not template_name:
        return
    elements = content.get("elements", {}) or {}
    input_signals = []
    for elem in elements.values():
        if elem.get("type") == "input-signal":
            signal_name = (elem.get("props") or {}).get("name")
            if signal_name:
                input_signals.append(signal_name)
    seen = set()
    ordered_inputs = []
    for name in input_signals:
        if name not in seen:
            ordered_inputs.append(name)
            seen.add(name)
    args_descriptions = project_meta.get("templateArgs") or {}
    args = {}
    arg_desc_lines = []
    for name in ordered_inputs:
        desc = (args_descriptions.get(name) or "").strip()
        arg_entry = {}
        if desc:
            arg_entry["description"] = desc
            arg_desc_lines.append(f"{name} - {desc}")
        else:
            arg_desc_lines.append(f"{name} -")
        args[name] = arg_entry
    general_description = (project_meta.get("description") or "").strip()
    full_description_parts = []
    if general_description:
        full_description_parts.append(general_description)
    if arg_desc_lines:
        full_description_parts.extend(arg_desc_lines)
    full_description = "; ".join(full_description_parts)
    entry = {
        "name": template_name,
        "args": args,
        "body": content.get("code", ""),
        "description": full_description
    }
    templates_data = load_templates()
    templates = templates_data.get("templates", [])
    found = False
    updated = []
    for tpl in templates:
        if tpl.get("name") == template_name:
            updated.append(entry)
            found = True
        else:
            updated.append(tpl)
    if not found:
        updated.append(entry)
    templates_data["templates"] = updated
    with open(TEMPLATES_PATH, "w", encoding="utf-8") as f:
        json.dump(templates_data, f, ensure_ascii=False, indent=2)
    STATE["templates"] = templates_data
    print(f"[OK] Template '{template_name}' saved/updated in formula_templates.json")

# =============================================================================
# FASTAPI ПРИЛОЖЕНИЕ
# =============================================================================

app = FastAPI(title="Logic Scheme Editor API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Хранилище сессий визуализатора
visualize_sessions: Dict[str, Dict[str, Any]] = {}

@app.on_event("startup")
def startup():
    settings = load_settings()
    STATE["settings"] = settings

        # Определяем список конфигураций по подпапкам в projects
    project_base = _abs_folder("projectDataFolder")
    configs = []
    if project_base and os.path.isdir(project_base):
        configs = sorted([
            d for d in os.listdir(project_base)
            if os.path.isdir(os.path.join(project_base, d))
        ])
    STATE["configurations"] = configs

    # Проверяем, изменились ли шаблоны (общие для всех конфигураций)
    templates_changed = False
    if os.path.exists(TEMPLATES_PATH):
        from update_projects import _file_hash, HASH_PATH
        new_hash = _file_hash(TEMPLATES_PATH)
        old_hash = None
        if os.path.exists(HASH_PATH):
            with open(HASH_PATH, "r", encoding="utf-8") as f:
                old_hash = f.read().strip()
        if new_hash != old_hash:
            templates_changed = True

    # Создаём подпапки и при необходимости обновляем проекты
    for config in configs:
        ensure_config_dirs(config)
        if templates_changed:
            proj_dir = config_path("projectDataFolder", config)
            if proj_dir:
                update_projects_if_templates_changed(
                    project_dir=proj_dir,
                    templates_path=TEMPLATES_PATH
                )

    # Сохраняем новый хэш только один раз после обработки всех конфигураций
    if templates_changed:
        with open(HASH_PATH, "w", encoding="utf-8") as f:
            f.write(new_hash)

    # Глобальные ресурсы
    STATE["templates"] = load_templates()
    os.makedirs(_abs_folder("templateDataFolder") or "", exist_ok=True)

    print(f"[OK] Configurations found: {configs}")
    print(f"[OK] Loaded templates: {len(STATE['templates'].get('templates', []))}")

# -----------------------------------------------------------------------------
# API: конфигурации
# -----------------------------------------------------------------------------

@app.get("/api/configurations")
def api_configurations():
    """Возвращает список доступных конфигураций."""
    return {"configurations": STATE["configurations"]}

# -----------------------------------------------------------------------------
# API: настройки и сигналы
# -----------------------------------------------------------------------------

@app.get("/api/settings")
def api_settings():
    return STATE["settings"]

@app.get("/api/tables")
def api_tables(q: str = "", limit: int = 50, config: str = Query(...)):
    items = get_tables_for_config(config)
    if not q:
        result_items = items[:limit]
        total = len(items)
    else:
        import re
        escaped = re.escape(q).replace(r"\*", ".*")
        rx = re.compile("^" + escaped + "$", re.IGNORECASE)
        filtered = [t for t in items if rx.match(t["Name"])]
        total = len(filtered)
        result_items = filtered[:max(1, min(limit, 500))]
    return JSONResponse(
        content={"items": result_items, "total": total},
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"}
    )

@app.post("/api/upload-signal-csv")
async def upload_signal_csv(file: UploadFile = File(...), config: str = Query(...)):
    """
    Сохраняет загруженный CSV‑файл в папку архива сигналов для заданной конфигурации.
    Файл должен содержать колонки DATE, TIME и хотя бы один столбец со значением сигнала.
    """
    archive_dir = config_path("signalArchiveFolder", config)
    if not archive_dir:
        raise HTTPException(status_code=500, detail="signalArchiveFolder not configured")
    os.makedirs(archive_dir, exist_ok=True)

    # Проверяем имя файла
    if not file.filename or '..' in file.filename or '/' in file.filename or '\\' in file.filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    destination = os.path.join(archive_dir, file.filename)
    with open(destination, "wb") as buffer:
        buffer.write(await file.read())

    return {"status": "ok", "filename": file.filename}

@app.get("/api/signals")
def api_signals(q: str = "", limit: int = 50, config: str = Query(...)):
    signals = get_signals_for_config(config)
    if not q:
        result = {"items": signals[:limit], "total": len(signals)}
    else:
        import re
        escaped = re.escape(q).replace(r"\*", ".*")
        rx = re.compile("^" + escaped + "$", re.IGNORECASE)
        items = [s for s in signals if rx.match(s["Tagname"])]
        result = {"items": items[:max(1, min(limit, 500))], "total": len(items)}
    return JSONResponse(
        content=result,
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"}
    )

@app.get("/api/table/file/{name}")
def api_table_file(name: str, config: str = Query(...)):
    folder = config_path("tablesFolder", config)
    if not folder:
        raise HTTPException(status_code=500, detail="tablesFolder not configured")
    if ".." in name or "/" in name or "\\" in name:
        raise HTTPException(status_code=400, detail="Invalid table name")
    path = os.path.join(folder, f"{name}.xlsx")
    if not path.startswith(folder):
        raise HTTPException(status_code=400, detail="Path traversal attempt")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Table file not found")
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{name}.xlsx",
        headers={"Cache-Control": "no-cache"}
    )

@app.get("/api/formula-templates")
def api_formula_templates():
    return STATE.get("templates") or {"templates": []}

# -----------------------------------------------------------------------------
# API: проекты
# -----------------------------------------------------------------------------

def collect_projects(directory: str, source_label: str) -> list[dict]:
    items = []
    if not directory or not os.path.isdir(directory):
        return items
    for fname in sorted(os.listdir(directory)):
        if not fname.endswith(".json"):
            continue
        path = os.path.join(directory, fname)
        try:
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except Exception:
            continue
        project_meta = payload.get("project", {}) or {}
        items.append({
            "filename": fname,
            "code": project_meta.get("code") or project_meta.get("tagname") or "",
            "description": project_meta.get("description") or "",
            "type": project_meta.get("type") or "",
            "source": source_label
        })
    return items

@app.get("/api/project/list")
def list_projects(config: str = Query(...)):
    def collect(directory_key: str, source_label: str) -> list[dict]:
        folder = config_path(directory_key, config) if directory_key == "projectDataFolder" else _abs_folder(directory_key)
        if not folder or not os.path.isdir(folder):
            return []
        items = []
        for fname in sorted(os.listdir(folder)):
            if not fname.endswith(".json"):
                continue
            path = os.path.join(folder, fname)
            try:
                with open(path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
            except Exception:
                continue
            project_meta = payload.get("project", {}) or {}
            try:
                code_len = get_code_length(payload)
            except Exception:
                code_len = 0
            items.append({
                "filename": fname,
                "code": project_meta.get("code") or project_meta.get("tagname") or "",
                "description": project_meta.get("description") or "",
                "type": project_meta.get("type") or "",
                "author": project_meta.get("author") or "",
                "possibleCause": project_meta.get("possibleCause") or "",
                "status": project_meta.get("status") or "draft",
                "statusComment": project_meta.get("statusComment") or "",
                "statusHistory": project_meta.get("statusHistory") or [],
                "lastStatusChangedByAdmin": project_meta.get("lastStatusChangedByAdmin", False),
                "lastModifiedAt": project_meta.get("lastModifiedAt") or "",
                "lastModifiedBy": project_meta.get("lastModifiedBy") or "",
                "source": source_label,
                "codeLength": code_len
            })
        return items

    projects = collect("projectDataFolder", "projects")
    projects.extend(collect("templateDataFolder", "templates"))
    authors = sorted(set(project.get("author") for project in projects if project.get("author")))
    admin_author = STATE["settings"].get("adminAuthor", "")
    return {"projects": projects, "authors": authors, "adminAuthor": admin_author}

@app.post("/api/project/set-author")
async def set_project_author(request: Request, config: str = Query(...)):
    try:
        data = await request.json()
        filename = data.get("filename")
        new_author = data.get("author")
        if not filename or not new_author:
            raise HTTPException(status_code=400, detail="Filename and author are required")
        path = get_storage_path(filename, storage="projects", config=config)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="Project not found")
        with open(path, "r", encoding="utf-8") as f:
            content = json.load(f)
        if "project" not in content:
            content["project"] = {}
        content["project"]["author"] = new_author
        with open(path, "w", encoding="utf-8") as f:
            json.dump(content, f, ensure_ascii=False, indent=2)
        invalidate_signals_cache(config)
        return {"status": "ok", "message": f"Author updated for {filename}"}
    except Exception as e:
        print(f"Error updating author: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/api/project/save")
async def save_project(request: Request, config: str = Query(...)):
    try:
        data = await request.json()
        filename = data.get("filename")
        content = data.get("content")
        target = data.get("target", "projects")
        if not filename or not content:
            raise HTTPException(status_code=400, detail="Filename and content are required")
        project_meta = content.get("project") or {}
        project_meta.setdefault("status", "draft")
        project_meta.setdefault("statusComment", "")
        project_meta.setdefault("statusHistory", [])
        project_type = project_meta.get("type", "parameter")
        if project_type == "template":
            target = "templates"
        path = get_storage_path(filename, storage=target, config=config if target == "projects" else None)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(content, f, ensure_ascii=False, indent=2)
        if project_type == "template":
            upsert_formula_template_from_project(content)
        invalidate_signals_cache(config)
        return {"status": "ok", "message": f"Project saved to {filename}"}
    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Error saving project: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during save")

@app.get("/api/project/load/{filename}")
def load_project(filename: str, source: str = "projects", config: str = Query(...)):
    try:
        path = get_storage_path(filename, storage=source, config=config if source == "projects" else None)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="Project not found")
        with open(path, "r", encoding="utf-8") as f:
            content = json.load(f)
        return content
    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Error loading project: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during load")

@app.post("/api/project/export-selected")
async def api_export_selected_projects(payload: dict = Body(...), config: str = Query(...)):
    try:
        filenames = payload.get("filenames", [])
        project_dir = config_path("projectDataFolder", config)
        if not project_dir:
            raise HTTPException(status_code=500, detail="projectDataFolder not configured")
        param_format = payload.get("param_format", "excel")
        result = export_selected_projects(filenames, project_dir, param_format)
        return StreamingResponse(
            BytesIO(result["content"]),
            media_type=result["media_type"],
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{quote(result['filename'])}",
                "Cache-Control": "no-cache, no-store, must-revalidate"
            }
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        print(f"Error exporting selected projects: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during export")

@app.get("/api/project/consumers/{code}")
def api_project_consumers(code: str, config: str = Query(...)):
    results = []
    # проекты в текущей конфигурации
    folder = config_path("projectDataFolder", config)
    if folder and os.path.isdir(folder):
        for fname in sorted(os.listdir(folder)):
            if not fname.endswith(".json"):
                continue
            path = os.path.join(folder, fname)
            try:
                with open(path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                proj_meta = (payload.get("project") or {})
                proj_code = (proj_meta.get("code") or proj_meta.get("tagname") or "").strip()
                elements = payload.get("elements") or {}
                inputs = []
                for el in elements.values():
                    if el and el.get("type") == "input-signal":
                        name = ((el.get("props") or {}).get("name") or "").strip()
                        if name:
                            inputs.append(name)
                if code in set(inputs):
                    results.append({
                        "code": proj_code or "(без кода)",
                        "filename": fname,
                        "source": "projects",
                        "description": proj_meta.get("description", ""),
                        "type": proj_meta.get("type", "")
                    })
            except Exception as e:
                print(f"[WARN] Failed to inspect project {path}: {e}")
                continue
    # шаблоны не привязаны к конфигурации, их тоже можно проверять
    tmpl_folder = _abs_folder("templateDataFolder")
    if tmpl_folder and os.path.isdir(tmpl_folder):
        for fname in sorted(os.listdir(tmpl_folder)):
            if not fname.endswith(".json"):
                continue
            path = os.path.join(tmpl_folder, fname)
            try:
                with open(path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                proj_meta = (payload.get("project") or {})
                proj_code = (proj_meta.get("code") or proj_meta.get("tagname") or "").strip()
                elements = payload.get("elements") or {}
                inputs = []
                for el in elements.values():
                    if el and el.get("type") == "input-signal":
                        name = ((el.get("props") or {}).get("name") or "").strip()
                        if name:
                            inputs.append(name)
                if code in set(inputs):
                    results.append({
                        "code": proj_code or "(без кода)",
                        "filename": fname,
                        "source": "templates",
                        "description": proj_meta.get("description", ""),
                        "type": proj_meta.get("type", "")
                    })
            except Exception as e:
                print(f"[WARN] Failed to inspect template {path}: {e}")
                continue
    results.sort(key=lambda x: x.get("code", "").lower())
    return {"consumers": results}

@app.post("/api/tags/list")
async def get_tags(config: str = Query(...)):
    # тэги глобальные, пока не зависят от конфигурации
    return load_tags_data()

@app.post("/api/tags/create")
async def create_tag(payload: dict = Body(...), config: str = Query(...)):
    user = payload.get("user")
    name = payload.get("name")
    color = payload.get("color")
    admin_author = STATE["settings"].get("adminAuthor", "")
    if user != admin_author:
        raise HTTPException(status_code=403, detail="Только админ может создавать теги")
    if not name or not color:
        raise HTTPException(status_code=400, detail="Нужно имя и цвет")
    data = load_tags_data()
    if any(t['name'] == name for t in data['tags']):
        raise HTTPException(status_code=400, detail="Тег с таким именем уже есть")
    new_tag = {"id": f"tag_{int(datetime.utcnow().timestamp())}", "name": name, "color": color}
    data["tags"].append(new_tag)
    save_tags_data(data)
    return new_tag

@app.post("/api/tags/assign")
async def assign_tags(payload: dict = Body(...)):   # убрали config
    filename = payload.get("filename")
    tag_ids = payload.get("tagIds", [])
    if not filename:
        raise HTTPException(status_code=400, detail="Нет имени файла")
    data = load_tags_data()
    data["assignments"][filename] = tag_ids
    save_tags_data(data)
    return {"status": "success"}

@app.post("/api/project/status/set")
async def set_project_status(payload: dict = Body(...), config: str = Query(...)):
    try:
        filename = payload.get("filename")
        new_status = payload.get("status")
        comment = payload.get("comment", "")
        user = payload.get("user", "")
        if not filename or new_status not in ("draft", "ready"):
            raise HTTPException(status_code=400, detail="Invalid filename or status")
        admin_author = STATE["settings"].get("adminAuthor", "")
        if user != admin_author:
            raise HTTPException(status_code=403, detail="Only admin can change status")
        file_path = get_storage_path(filename, storage="projects", config=config)
        if not os.path.isfile(file_path):
            raise HTTPException(status_code=404, detail=f"Project file not found: {filename}")
        with open(file_path, "r", encoding="utf-8") as f:
            content = json.load(f)
        project_meta = content.setdefault("project", {})
        project_meta["status"] = new_status
        project_meta["lastStatusChangedByAdmin"] = True
        history = project_meta.setdefault("statusHistory", [])
        history.append({
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "author": user,
            "action": "forced_draft" if new_status == "draft" else "forced_ready",
            "comment": comment.strip()
        })
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(content, f, ensure_ascii=False, indent=2)
        return {"status": "success", "message": "Project status updated"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating project status: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.delete("/api/project/delete/{filename}")
async def delete_project(filename: str, request: Request, config: str = Query(...)):
    try:
        data = await request.json()
        current_user = data.get("user", "")
        if not current_user:
            raise HTTPException(status_code=400, detail="User not provided")
        path = get_storage_path(filename, storage="projects", config=config)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="Project not found")
        with open(path, "r", encoding="utf-8") as f:
            content = json.load(f)
        project_meta = content.get("project", {})
        project_author = project_meta.get("author", "")
        admin_author = STATE["settings"].get("adminAuthor", "")
        if project_author != current_user and current_user != admin_author:
            raise HTTPException(status_code=403, detail="Only the author or admin can delete this project")
        # Перемещаем в deleted_projects/<config>
        deleted_base = _abs_folder("deletedFolder") or os.path.join(BASE_DIR, "deleted_projects")
        deleted_dir = os.path.join(deleted_base, config)
        os.makedirs(deleted_dir, exist_ok=True)
        deleted_path = os.path.join(deleted_dir, filename)
        shutil.copy2(path, deleted_path)
        os.remove(path)
        invalidate_signals_cache(config)
        return {"status": "ok", "message": f"Project '{filename}' successfully moved to deleted_projects"}
    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Error deleting project: {e}")
        raise HTTPException(status_code=500, detail="Internal server error during deletion")

# -----------------------------------------------------------------------------
# API: данные сигналов и зависимости
# -----------------------------------------------------------------------------

@app.post("/api/signal-data")
async def api_signal_data(request: Request, config: str = Query(...)):
    try:
        data = await request.json()
        signal_names = data.get("signal_names", [])
        output_format = data.get("format", "parquet")
        if not signal_names:
            raise HTTPException(status_code=400, detail="signal_names is required")
        signals_data = load_signal_data_optimized(signal_names, config)
        response = {
            "found": list(signals_data.keys()),
            "not_found": [s for s in signal_names if s not in signals_data],
            "format": output_format
        }
        if not signals_data:
            raise HTTPException(status_code=404, detail="No signals found")
        if output_format == "parquet":
            return await _export_parquet(signals_data, response)
        else:
            return await _export_json(signals_data, response)
    except HTTPException as e:
        raise e
    except Exception as e:
        print(f"Error in api_signal_data: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def _export_parquet(signals_data: Dict[str, pd.DataFrame], meta: Dict):
    try:
        with tempfile.NamedTemporaryFile(suffix=".parquet", delete=False) as tmp:
            tmp_path = tmp.name
        rows = []
        for signal_name, df in signals_data.items():
            df_copy = df.copy()
            df_copy["signal_name"] = signal_name
            rows.append(df_copy)
        combined = pd.concat(rows, ignore_index=True)
        combined.to_parquet(tmp_path, compression='snappy', index=False)
        file_size = os.path.getsize(tmp_path)
        print(f"[OK] Exported {len(signals_data)} signals to Parquet: {file_size / 1024 / 1024:.2f} MB")
        return FileResponse(
            tmp_path,
            media_type="application/octet-stream",
            filename="signal_data.parquet",
            headers={"X-Signal-Meta": json.dumps(meta)}
        )
    except Exception as e:
        print(f"[ERROR] Parquet export failed: {e}")
        raise

async def _export_json(signals_data: Dict[str, pd.DataFrame], meta: Dict):
    try:
        data_dict = {}
        for signal_name, df in signals_data.items():
            df_copy = df.copy()
            df_copy["datetime"] = df_copy["datetime"].astype(str)
            data_dict[signal_name] = df_copy.to_dict(orient="records")
        response_data = {**meta, "data": data_dict}
        return JSONResponse(response_data)
    except Exception as e:
        print(f"[ERROR] JSON export failed: {e}")
        raise

@app.post("/api/resolve-signals")
async def api_resolve_signals(request: Request, config: str = Query(...)):
    try:
        data = await request.json()
        signal_names = data.get("signals", [])
        base_signals, synthetic_signals = resolve_signal_dependencies(signal_names, config)
        computation_order = topological_sort_signals(synthetic_signals)
        return {
            "base_signals": list(base_signals),
            "synthetic_signals": synthetic_signals,
            "computation_order": computation_order
        }
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        print(f"[ERROR] resolve-signals failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# -----------------------------------------------------------------------------
# API: визуализатор (без изменений, не зависит от файловой структуры)
# -----------------------------------------------------------------------------

@app.post("/api/visualize/session")
async def create_visualize_session(request: Request):
    try:
        data = await request.json()
        signals = data.get("signals", [])
        tables = data.get("tables", [])
        code = data.get("code", "")
        visualizer_state = data.get("visualizer_state")
        if not isinstance(signals, list):
            raise HTTPException(status_code=400, detail="signals must be a list")
        token = uuid.uuid4().hex
        visualize_sessions[token] = {
            "signals": signals,
            "tables": tables,
            "code": code,
            "visualizer_state": visualizer_state
        }
        return {"token": token}
    except Exception as e:
        print(f"[ERROR] create_visualize_session failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/visualize/session/{token}")
async def get_visualize_session(token: str):
    session = visualize_sessions.get(token)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return {
        "signals": session.get("signals", []),
        "tables": session.get("tables", []),
        "code": session.get("code", ""),
        "visualizer_state": session.get("visualizer_state")
    }

@app.post("/api/visualize/save-state")
async def save_visualizer_state(request: VisualizerStateRequest) -> VisualizerStateResponse:
    try:
        if request.session_token in visualize_sessions:
            visualize_sessions[request.session_token]["visualizer_state"] = request.state
        else:
            visualize_sessions[request.session_token] = {
                "signals": [],
                "code": "",
                "visualizer_state": request.state
            }
        return VisualizerStateResponse(success=True, state=request.state, message="Состояние сохранено")
    except Exception as e:
        return VisualizerStateResponse(success=False, message=f"Ошибка сохранения: {str(e)}")

@app.get("/api/visualize/get-state/{session_token}")
async def get_visualizer_state(session_token: str) -> VisualizerStateResponse:
    session = visualize_sessions.get(session_token)
    if session is None:
        return VisualizerStateResponse(success=False, message="Сессия не найдена")
    state = session.get("visualizer_state")
    if state is None:
        return VisualizerStateResponse(success=False, message="Состояние визуализатора не сохранено")
    return VisualizerStateResponse(success=True, state=state)

@app.get("/api/tags/list")
async def get_tags_list():
    return load_tags_data()


@app.post("/api/check-syntax")
async def check_syntax(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Поддерживаются только файлы Excel (.xlsx, .xls)")

    contents = await file.read()

    # Читаем Excel вручную, чтобы избежать ошибок приведения типов
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Пустой лист")

        # Первая строка — заголовки
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        data = []
        for row in rows[1:]:
            data.append([str(cell) if cell is not None else "" for cell in row])
        df = pd.DataFrame(data, columns=headers)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel: {e}")

    code_col = next((c for c in df.columns if c.lower().strip() in ['код', 'code']), None)

    code_col = next((c for c in df.columns if c.lower().strip() in ['код', 'code']), None)
    signals_col = next((c for c in df.columns if c.lower().strip() in ['используемые сигналы', 'used signals']), None)

    if not code_col:
        raise HTTPException(status_code=400, detail="В файле не найден столбец 'Код'")
    if not signals_col:
        raise HTTPException(status_code=400, detail="В файле не найден столбец 'Используемые сигналы'")

    remarks = []
    for idx, row in df.iterrows():
        code = str(row[code_col]) if pd.notna(row[code_col]) else ""
        signals_str = str(row[signals_col]) if pd.notna(row[signals_col]) else ""
        input_signals = [s.strip() for s in re.split(r'[;,\n]', signals_str) if s.strip()]

        row_remarks = []
        row_num = idx + 2

        if not code:
            row_remarks.append("Пустой код")
            remarks.append({"row": row_num, "remarks": row_remarks})
            continue

        # 0. Проверка имён сигналов на недопустимые символы (арифметические операторы)
        for sig in input_signals:
            if any(c in sig for c in "+-*/^"):
                row_remarks.append(f"Сигнал '{sig}' содержит недопустимый символ в имени (возможно, пропущен оператор)")

        # Проверка, является ли код просто числом или прочерком
        is_constant = False
        clean_code = code.strip()
        if clean_code and (clean_code.replace('.', '', 1).replace(',', '', 1).isdigit() or clean_code == '-'):
            is_constant = True

        # 1. Скобки и кавычки
        if code.count('(') != code.count(')'):
            row_remarks.append("Не совпадает количество открывающих и закрывающих скобок")
        if code.count("'") % 2 != 0:
            row_remarks.append("Нечётное количество одинарных кавычек")
        if code.count('"') % 2 != 0:
            row_remarks.append("Нечётное количество двойных кавычек")

        # 2. Недопустимые символы (кириллица запрещена)
        allowed_chars = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_.,;:!?+-*/%<>=&|^()[]{}§'\"\n\r\t ")
        invalid_chars = set(code) - allowed_chars
        if invalid_chars:
            row_remarks.append(f"Недопустимые символы: {', '.join(repr(c) for c in invalid_chars)}")

                # 2b. Проверка на экспоненциальную запись числа (недопустимо)
        if re.search(r'\b\d+\.?\d*[eE][+-]?\d+\b', code):
            row_remarks.append("Обнаружено число в экспоненциальной записи (например, 1.5E-3). Такая запись недопустима – используйте десятичную дробь.")

        # 3. Унарный минус перед идентификатором (не перед числом и не перед скобкой)
        if not is_constant and re.search(r'(?<![a-zA-Z0-9_§])\s*-\s*(?=[A-Za-z_§])', code):
            row_remarks.append("Обнаружен унарный минус перед сигналом. При необходимости замените на '-1*'")

        # 4. Логические операторы
        for op in ['AND', 'OR', 'NOT']:
            if re.search(rf'\b{op}\b', code):
                row_remarks.append(f"Логический оператор {op} – рекомендуется заменить на {'&&' if op=='AND' else '||' if op=='OR' else '!'}")

               # 4b. Одиночные & и | (не являются частью && или ||)
        if re.search(r'(?<![&])&(?![&])', code):
            row_remarks.append("Обнаружен одиночный '&'. Возможно, вы имели в виду '&&' (логическое И).")
        if re.search(r'(?<![|])\|(?![|])', code):
            row_remarks.append("Обнаружен одиночный '|'. Возможно, вы имели в виду '||' (логическое ИЛИ).")
        # 4c. Проверка одиночного '=' (вместо '==')
        code_no_strings = re.sub(r'[\'"].*?[\'"]', '', code)
        if re.search(r'(?<![<>=!])=(?![=])', code_no_strings):
                row_remarks.append("Обнаружен одиночный '='. Возможно, вы имели в виду '==' (сравнение).")

         # 5. Аргументы HISTORY*/PREV – всегда должны быть в кавычках
        history_funcs = ['HISTORYAVG','HISTORYCOUNT','HISTORYSUM','HISTORYMAX','HISTORYMIN','HISTORYDIFF','HISTORYGRADIENT','PREV']
        for fn in history_funcs:
            pattern = re.compile(rf'\b{fn}\s*\(\s*([\'"]?)(?P<arg>[^\'",]+)\1\s*[,)]', re.IGNORECASE)
            for m in pattern.finditer(code):
                arg = m.group('arg').strip()
                if arg and (m.group(1) is None or m.group(1) == ''):
                    row_remarks.append(f"{fn}: аргумент '{arg}' должен быть в кавычках (обязательно для {fn})")

        # 5b. Аргументы INTERPOLATE/GETPOINT – всегда должны быть в кавычках
        for fn in ['INTERPOLATE', 'GETPOINT']:
            pattern = re.compile(rf'\b{fn}\s*\(\s*([\'"]?)(?P<arg>[^\'",]+)\1\s*[,)]', re.IGNORECASE)
            for m in pattern.finditer(code):
                arg = m.group('arg').strip()
                if arg and (m.group(1) is None or m.group(1) == ''):
                    row_remarks.append(f"{fn}: аргумент '{arg}' должен быть в кавычках (обязательно для {fn})")

        # 6. Неизвестные функции/опечатки
        if not is_constant:
            known_functions = {'WHEN','ABS','EXP','POW','LOG','LOG10','MIN','MAX','AVG','MED','ROUND',
                            'GETPOINT','INTERPOLATE','PREV','HISTORYAVG','HISTORYCOUNT','HISTORYSUM',
                            'HISTORYMAX','HISTORYMIN','HISTORYDIFF','HISTORYGRADIENT'}

            string_literals = re.findall(r'[\'"].*?[\'"]', code)

            for token in re.findall(r'[A-Z][A-Z0-9_]*', code):
                if token in known_functions or token in ('AND','OR','NOT','X','Y'):
                    continue
                if token in input_signals:
                    continue
                if re.match(r'P\d', token):
                    continue
                if any(token in s for s in string_literals):
                    continue
                if any(token in sig.replace('§', '_') for sig in input_signals):
                    continue
                row_remarks.append(f"Возможно, неизвестная функция или опечатка: '{token}'")

        # 7. Разрыв кода (пробел между идентификаторами)
        if not is_constant:
            code_no_strings = re.sub(r'[\'"].*?[\'"]', '', code)
            if re.search(r'[A-Za-z0-9_§]+\s+[A-Za-z0-9_§]+', code_no_strings):
                row_remarks.append("Обнаружен разрыв кода: два идентификатора или значение через пробел")

        # 8. Проверка использования сигналов (пропускается для констант)
        if not is_constant:
            for sig in input_signals:
                sig_underscored = sig.replace('§', '_')
                found = (sig in code) or (sig_underscored in code)
                if not found and sig[0].isdigit():
                    found = ('P' + sig in code) or ('P' + sig_underscored in code)
                if not found:
                    row_remarks.append(f"Сигнал '{sig}' не найден в выражении")

        # 9. Префикс P для сигналов, начинающихся с цифры
        if not is_constant:
            for sig in input_signals:
                if not sig[0].isdigit():
                    continue
                sig_u = sig.replace('§', '_')
                code_without_strings = re.sub(r'[\'"].*?[\'"]', '', code)
                pattern_sig = re.compile(rf'(?<![A-Za-z0-9_]){re.escape(sig_u)}(?![A-Za-z0-9_])')
                if pattern_sig.search(code_without_strings):
                    row_remarks.append(f"Сигнал '{sig}' начинается с цифры – необходимо добавить префикс P")

        # 10. Проверка "голых" сигналов в логических условиях
        if not is_constant:
            code_no_strings = re.sub(r'[\'"].*?[\'"]', '', code)
            for sig in input_signals:
                sig_u = sig.replace('§', '_')
                # Возможные варианты написания сигнала (с учётом префикса P для цифровых)
                candidates = {sig, sig_u}
                if sig[0].isdigit():
                    candidates.add('P' + sig)
                    candidates.add('P' + sig_u)

                for cand in candidates:
                    # Ищем вхождения вида (сигнал) с возможными пробелами
                    pattern = re.compile(rf'\(\s*({re.escape(cand)})\s*\)')
                    for m in pattern.finditer(code_no_strings):
                        token = m.group(1)
                        # Проверяем, не является ли это аргументом функции:
                        # перед скобкой не должно быть имени функции/сигнала
                        before = code_no_strings[:m.start()]
                        if re.search(r'[A-Za-z0-9_§]\s*$', before):
                            continue
                        row_remarks.append(
                            f"Сигнал '{token}' используется как логическое условие без сравнения или оператора. "
                            "Возможно, пропущен оператор (например, = 1)."
                        )
                        break  # для каждого сигнала достаточно одного замечания

        if row_remarks:
            remarks.append({"row": row_num, "remarks": row_remarks})

        

    return remarks

@app.post("/api/export-remarks")
async def export_remarks(data: List[Dict[str, Any]] = Body(...)):
    """Принимает массив замечаний и возвращает Excel-файл."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Замечания"
    ws.append(["Строка", "Замечания"])
    for item in data:
        remarks_text = "\n".join(item.get("remarks", []))
        ws.append([item.get("row", ""), remarks_text])
    # Автоширина и перенос текста
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    max_line = max(len(line) for line in lines)
                    if max_line > max_length:
                        max_length = max_line
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[col_letter].width = adjusted_width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=syntax_remarks.xlsx"}
    )



# Зависимости: используем уже существующие функции
# – extract_input_signals_from_project, resolve_signal_dependencies,
# – load_project_by_code, config_path

def _build_dependency_tree(signal_name: str, config: str, visited: set = None) -> dict:
    if visited is None:
        visited = set()
    if signal_name in visited:
        return {"name": signal_name, "type": "cyclic", "description": "Циклическая зависимость", "code": ""}
    visited.add(signal_name)

    # Базовый сигнал
    if is_base_signal(signal_name, config):
        # Ищем описание среди загруженных сигналов
        desc = ""
        signals_list = get_signals_for_config(config)
        for s in signals_list:
            if s["Tagname"] == signal_name:
                desc = s.get("Description", "")
                break
        return {
            "name": signal_name,
            "type": "base",
            "description": desc,
            "code": ""   # у базового сигнала нет кода
        }

    # Синтетический сигнал – загружаем его проект
    project = load_project_by_code(signal_name, config)
    if not project:
        return {"name": signal_name, "type": "unknown", "description": "Проект не найден", "code": ""}

    proj_meta = project.get("project", {})
    desc = proj_meta.get("description", "")
    dim = proj_meta.get("dimension", "")
    code = project.get("formula", "")   # исходный код проекта
    code_truncated = False
    if len(code) > STATE.get('settings').get('llm').get('max_code_length'):
        code = code[:STATE.get('settings').get('llm').get('max_code_length')] + "..."
        code_truncated = True

    # Рекурсивно строим детей
    input_names = extract_input_signals_from_project(project)
    children = []
    for inp in sorted(set(input_names)):
        children.append(_build_dependency_tree(inp, config, visited.copy()))

    return {
        "name": signal_name,
        "type": "synthetic",
        "description": desc,
        "dimension": dim,
        "code": code,
        "code_truncated": code_truncated,
        "inputs": children
    }


@app.get("/api/project/dependency-tree")
def api_dependency_tree(filename: str, source: str = "projects", config: str = Query(...)):
    """Возвращает дерево зависимостей сигналов для указанного проекта."""
    # Загружаем проект
    try:
        path = get_storage_path(filename, storage=source, config=config if source == "projects" else None)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="Project not found")
        with open(path, "r", encoding="utf-8") as f:
            content = json.load(f)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading project: {e}")

    project_meta = content.get("project", {})
    project_code = project_meta.get("code", "").strip()
    if not project_code:
        raise HTTPException(status_code=400, detail="Project code not found")

    # Собираем входные сигналы текущего проекта
    input_signals = extract_input_signals_from_project(content)
    unique_inputs = sorted(set(input_signals))

    # Строим дерево для каждого входного сигнала
    trees = []
    for sig in unique_inputs:
        trees.append(_build_dependency_tree(sig, config))

    raw_code = content.get("code", "")
    code_truncated = False
    if len(raw_code) > STATE.get('settings').get('llm').get('max_code_length'):
        raw_code = raw_code[:STATE.get('settings').get('llm').get('max_code_length')] + "..."
        code_truncated = True

    return {
        "project": project_code,
        "type": project_meta.get("type", ""),
        "description": project_meta.get("description", ""),
        "dimension": project_meta.get("dimension", ""),
        "possibleCause": project_meta.get("possibleCause", ""),
        "guidelines": project_meta.get("guidelines", ""),
        "code": raw_code,          # <-- код текущего проекта
        "code_truncated": code_truncated,
        "dependencies": trees
    }

@app.get("/api/llm/context-structure")
def api_llm_context_structure():
    """Возвращает содержимое structure.md из папки, указанной в настройках"""
    settings = STATE.get("settings") or {}
    llm_cfg = settings.get("llm") or {}
    context_dir = llm_cfg.get("contextFolder")
    if not context_dir:
        raise HTTPException(status_code=500, detail="LLM context folder not configured in settings.json")
    # Если путь относительный – делаем абсолютным относительно BASE_DIR
    if not os.path.isabs(context_dir):
        context_dir = os.path.normpath(os.path.join(BASE_DIR, context_dir))
    path = os.path.join(context_dir, "structure.md")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="structure.md not found")
    with open(path, "r", encoding="utf-8") as f:
        return {"content": f.read()}


@app.get("/api/llm/context-syntax")
def api_llm_context_syntax():
    """Возвращает содержимое syntax.md из папки, указанной в настройках"""
    settings = STATE.get("settings") or {}
    llm_cfg = settings.get("llm") or {}
    context_dir = llm_cfg.get("contextFolder")
    if not context_dir:
        raise HTTPException(status_code=500, detail="LLM context folder not configured in settings.json")
    if not os.path.isabs(context_dir):
        context_dir = os.path.normpath(os.path.join(BASE_DIR, context_dir))
    path = os.path.join(context_dir, "syntax.md")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="syntax.md not found")
    with open(path, "r", encoding="utf-8") as f:
        return {"content": f.read()}
    
@app.get("/api/llm/config")
def api_llm_config():
    settings = STATE.get("settings") or {}
    llm_cfg = settings.get("llm") or {}
    return {
        "ollamaUrl": llm_cfg.get("ollamaUrl", "http://localhost:11434"),
        "contextFolder": llm_cfg.get("contextFolder", ""),
        "max_code_length": llm_cfg.get("max_code_length", 4000),
        "model": llm_cfg.get("model", "gemma4:31b")   # ← новое
    }

@app.post("/api/llm/generate")
async def api_llm_generate(payload: dict = Body(...)):
    settings = STATE.get("settings") or {}
    llm_cfg = settings.get("llm") or {}
    ollama_url = llm_cfg.get("ollamaUrl", "http://localhost:11434")
    model = llm_cfg.get("model", "gemma4:31b")
    prompt = payload.get("prompt", "")
    timeout = llm_cfg.get("timeout", 300)

    if not prompt:
        raise HTTPException(status_code=400, detail="Prompt is required")

    async with httpx.AsyncClient(timeout=timeout) as client:
        try:
            resp = await client.post(
                f"{ollama_url}/api/generate",
                json={"model": model, "prompt": prompt, "stream": False}
            )
            resp.raise_for_status()
            data = resp.json()
            return {"response": data.get("response", "")}
        except httpx.HTTPStatusError as e:
            raise HTTPException(status_code=502, detail=f"Ollama HTTP error: {e.response.text}")
        except httpx.TimeoutException:
            raise HTTPException(status_code=504, detail="Ollama request timed out")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Ollama connection error: {str(e)}")
        

@app.get("/api/nn-block-params")
async def get_nn_block_params():
    json_path = os.path.join(BASE_DIR, "nn_block_params.json")
    if not os.path.exists(json_path):
        raise HTTPException(status_code=404, detail="nn_block_params.json not found")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Поддерживаем и массив, и объект: если объект, отдаём список значений
    if isinstance(data, dict):
        data = list(data.values())
    return JSONResponse(content=data)

@app.post("/api/nn/apply")
async def apply_nn_processing(payload: dict = Body(...)):
    try:
        element_id = payload["element_id"]
        project = payload["project"]
        config = payload.get("config") or ""
        project_code = project.get("project", {}).get("code", "unnamed_project")
        result = process_element(element_id, project, config, project_code)
        return {"status": "success", **result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

@app.delete("/api/nn/data/{element_id}")
async def delete_nn_data(element_id: str, config: str = Query(...), code: str = Query(...)):
    meta_path = get_meta_path(config, code, element_id)
    paths_to_delete = []
    had_meta = os.path.exists(meta_path)
    if had_meta:
        with open(meta_path, 'r') as f:
            meta = json.load(f)
        outputs = meta.get('outputs', {})
        for port, rel_path in outputs.items():
            if rel_path:
                paths_to_delete.append(os.path.join(get_datasets_dir(config), rel_path))
    # Также удаляем основной файл (на случай, если outputs нет)
    main_path = get_dataset_path(config, code, element_id)
    paths_to_delete.append(main_path)
    paths_to_delete.append(meta_path)
    # Добавляем возможные старые файлы _out0, _out1, _X, _y для совместимости
    for suffix in ['_out0', '_out1', '_X', '_y']:
        p = get_dataset_path(config, code, element_id + suffix)
        paths_to_delete.append(p)
        mp = get_meta_path(config, code, element_id + suffix)
        paths_to_delete.append(mp)

    deleted, not_found, errors = [], [], []
    for path in set(paths_to_delete):
        if os.path.exists(path):
            try:
                os.remove(path)
                deleted.append(path)
            except Exception as e:
                errors.append({"path": path, "error": str(e)})
        else:
            not_found.append(path)

    if not had_meta and not deleted:
        # Ничего не нашли даже по meta.json — почти всегда это означает,
        # что config/code при удалении не совпадает с тем, что был при
        # сохранении (переименование проекта/смена конфига), либо элемент
        # никогда не применялся.
        print(f"[delete_nn_data] Нет данных для {element_id} по пути {meta_path} "
              f"(config={config!r}, code={code!r}). Проверьте, не менялись ли "
              f"название проекта/конфиг после применения этого элемента.")

    return {
        "status": "deleted" if deleted else "nothing_found",
        "deleted": deleted,
        "not_found": not_found,
        "errors": errors,
    }
    
@app.get("/api/nn/data/{element_id}/columns")
async def get_dataset_columns(element_id: str, config: str = Query(...), code: str = Query(...), port: str = Query('out-0')):
    from dataprocessing import get_output_file
    elem = {'id': element_id}
    try:
        data_path = get_output_file(elem, port, config, code)
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))
    if not os.path.exists(data_path):
        raise HTTPException(status_code=404, detail="File not found")
    df = pd.read_excel(data_path, nrows=0)
    columns = [c for c in df.columns if c != 'datetime']
    return {"columns": columns}

@app.get("/api/nn/data/{element_id}/timerange")
async def get_dataset_timerange(element_id: str, config: str = Query(...), code: str = Query(...), port: str = Query('out-0')):
    elem = {'id': element_id}
    data_path = get_output_file(elem, port, config, code)
    if not os.path.exists(data_path):
        raise HTTPException(status_code=404, detail="File not found")
    df = pd.read_excel(data_path)
    df['datetime'] = pd.to_datetime(df['datetime'])
    return {"min_date": df['datetime'].min().strftime('%Y-%m-%dT%H:%M:%S'),
            "max_date": df['datetime'].max().strftime('%Y-%m-%dT%H:%M:%S')}

@app.get("/api/nn/data/{element_id}/stats")
async def get_dataset_stats(element_id: str, config: str = Query(...), code: str = Query(...), port: str = Query('out-0')):
    elem = {'id': element_id}
    data_path = get_output_file(elem, port, config, code)
    if not os.path.exists(data_path):
        raise HTTPException(status_code=404, detail="File not found")
    df = pd.read_excel(data_path)
    stats = {}
    for col in df.columns:
        if col == 'datetime': continue
        numeric_col = pd.to_numeric(df[col], errors='coerce').dropna()
        if not numeric_col.empty:
            stats[col] = {'min': float(numeric_col.min()), 'max': float(numeric_col.max()),
                          'mean': float(numeric_col.mean()), 'median': float(numeric_col.median())}
    return {"columns": stats}
    return {"columns": stats}
    
@app.get("/api/nn/data/{element_id}/stats")
async def get_dataset_stats(element_id: str, config: str = Query(...), code: str = Query(...)):
    data_path = get_dataset_path(config, code, element_id)
    if not os.path.exists(data_path):
        raise HTTPException(status_code=404, detail="Dataset file not found")
    try:
        df = pd.read_excel(data_path)
        stats = {}
        for col in df.columns:
            if col == 'datetime':
                continue
            numeric_col = pd.to_numeric(df[col], errors='coerce').dropna()
            if not numeric_col.empty:
                stats[col] = {
                    'min': float(numeric_col.min()),
                    'max': float(numeric_col.max()),
                    'mean': float(numeric_col.mean()),
                    'median': float(numeric_col.median())
                }
        return {"columns": stats}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.get("/api/nn/data/{element_id}/timerange")
async def get_dataset_timerange(element_id: str, config: str = Query(...), code: str = Query(...), port: str = Query('out-0')):
    elem = {'id': element_id}
    data_path = get_output_file(elem, port, config, code)
    if not os.path.exists(data_path):
        raise HTTPException(status_code=404, detail="File not found")
    df = pd.read_excel(data_path)
    df['datetime'] = pd.to_datetime(df['datetime'])
    return {"min_date": df['datetime'].min().strftime('%Y-%m-%dT%H:%M:%S'),
            "max_date": df['datetime'].max().strftime('%Y-%m-%dT%H:%M:%S')}

@app.post("/api/nn/status/{element_id}")
async def get_nn_status(element_id: str, payload: dict = Body(...)):
    """
    Лёгкая проверка актуальности данных элемента (без записи файлов).
    Используется фронтом при каждом открытии модалки свойств,
    чтобы не полагаться на локальный client-side кэш.
    """
    try:
        project = payload["project"]
        config = payload.get("config") or ""
        project_code = project.get("project", {}).get("code", "unnamed_project")
        status = get_element_status(element_id, project, config, project_code)
        return status
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.get("/api/nn/list")
async def list_processed_elements(config: str = Query(...), code: str = Query(...)):
    """Возвращает элементы проекта с существующими файлами данных."""
    datasets_dir = get_datasets_dir(config)
    if not os.path.exists(datasets_dir):
        return {"elements": []}
    elements = []
    prefix = code + "_"
    for fname in os.listdir(datasets_dir):
        if fname.startswith(prefix) and fname.endswith(".xlsx"):
            element_id = fname[len(prefix):-5]  # убираем префикс и .xlsx
            meta_path = os.path.join(datasets_dir, f"{code}_{element_id}_meta.json")
            description = ""
            if os.path.exists(meta_path):
                with open(meta_path, 'r') as f:
                    meta = json.load(f)
                description = meta.get('hash', '')[:8]  # короткий хэш для идентификации
            elements.append({
                "element_id": element_id,
                "filename": fname,
                "description": description
            })
    return {"elements": elements}

@app.get("/api/nn/data/{element_id}/full")
async def get_full_dataset(element_id: str, config: str = Query(...), code: str = Query(...), port: str = Query('out-0')):
    """Возвращает полный DataFrame в виде JSON."""
    from dataprocessing import get_output_file
    elem = {'id': element_id}
    try:
        data_path = get_output_file(elem, port, config, code)
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))
    if not os.path.exists(data_path):
        raise HTTPException(status_code=404, detail="File not found")
    df = pd.read_excel(data_path)
    # Преобразуем datetime в строку для JSON
    if 'datetime' in df.columns:
        df['datetime'] = df['datetime'].astype(str)
    return JSONResponse(content=json.loads(df.to_json(orient='records', date_format='iso')))





@app.on_event("startup")
async def _start_training_dispatcher():
    nn_template.set_config_path(config_path)
    # Если design-проекты сети у вас сохраняются с project.type == "template" —
    # main.save_project() уводит их в глобальную templateDataFolder, а не в
    # projectDataFolder/<config>. Эта строка добавляет её в область поиска
    # list_design_projects()/load_design_project(). Если у вас всё лежит в
    # projectDataFolder — эта строка ничего не сломает, просто добавит ещё
    # один (скорее всего пустой для дизайнов) каталог в скан.
    nn_template.set_template_dir(lambda: _abs_folder("templateDataFolder"))
    tq.start_dispatcher()
 
# Диагностика: если после этого поиск всё ещё пустой — временно включите
# подробный лог сканирования файлов проектов (что нашли, какие у них
# project.type / element_types / is_design), запустив сервер с переменной
# окружения NN_TEMPLATE_DEBUG=1, например:
#     NN_TEMPLATE_DEBUG=1 uvicorn main:app ...
# В логе появится по одной строке на каждый просканированный .json —
# сразу будет видно, находит ли скрипт файл дизайна вообще, и если да —
# почему is_design_project() возвращает False.
 
 
# ---------------------------------------------------------------------------
# Автокомплит: список кодов (ККС) дизайн-проектов сети для поля в модалке
# ---------------------------------------------------------------------------
 
@app.get("/api/nn/designs")
async def list_nn_designs(config: str = Query(...), query: str = Query('')):
    try:
        codes = nn_template.list_design_projects(config, query)
        return {"designs": codes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
 
 
# ВРЕМЕННЫЙ debug-эндпоинт — удалить после того, как разберёмся с поиском
# дизайнов. Открыть в браузере: /api/nn/designs/debug?config=TEC_20
# Покажет прямо в теле ответа: реальный абсолютный путь к папке проектов,
# какие .json файлы там есть, code/type/element_types каждого, и почему
# is_design_project() принял или отклонил файл.
@app.get("/api/nn/designs/debug")
async def debug_nn_designs(config: str = Query(...)):
    return nn_template.debug_scan(config)
 
 
# ---------------------------------------------------------------------------
# Статус элемента "Шаблон" (обучен / актуален / идёт обучение) —
# дёргается фронтом при открытии проекта и при открытии модалки свойств
# ---------------------------------------------------------------------------
 
@app.post("/api/nn/template/status/{element_id}")
async def get_template_status(element_id: str, payload: dict = Body(...)):
    try:
        project = payload["project"]
        config = payload.get("config") or ""
        project_code = project.get("project", {}).get("code", "unnamed_project")
 
        def _job_lookup(el_id, proj_code):
            return tq.find_active_job_for_element(el_id, proj_code)
 
        status = nn_template.get_template_status(element_id, project, config, project_code,
                                                   current_job_lookup=_job_lookup)
        return status
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
 
 
# ---------------------------------------------------------------------------
# Постановка обучения в очередь
# ---------------------------------------------------------------------------
 
# Порты элемента "Шаблон" в порядке, зафиксированном на фронте (см. neural_app.js):
# in-0 = settings, in-1 = X_train, in-2 = Y_train, in-3 = X_val, in-4 = Y_val
TEMPLATE_PORT_MAP = {
    'in-0': 'settings',
    'in-1': 'X_train',
    'in-2': 'Y_train',
    'in-3': 'X_val',
    'in-4': 'Y_val',
}
 
 
@app.post("/api/nn/train")
async def start_training(payload: dict = Body(...)):
    """
    payload = {
        "element_id": ...,
        "project": {...},          # как и в остальных nn-эндпоинтах
        "config": ...,
        "user": "имя пользователя" # из AppState.currentUser на фронте
    }
    """
    try:
        element_id = payload["element_id"]
        project = payload["project"]
        config = payload.get("config") or ""
        user = payload.get("user") or "Аноним"
        project_code = project.get("project", {}).get("code", "")
 
        # 1. Проект должен быть сохранён — иначе после закрытия вкладки
        #    некуда будет вернуться проверять статус обучения.
        if not project_code:
            raise HTTPException(status_code=400, detail="Сначала сохраните проект (укажите код)")
 
        elements = project.get("elements", {})
        elem = elements.get(element_id)
        if not elem:
            raise HTTPException(status_code=404, detail="Элемент не найден")
 
        design_code = elem.get("props", {}).get("design_code", "").strip()
        if not design_code:
            raise HTTPException(status_code=400, detail="Не указан код проекта дизайна нейросети")
 
        # 2. Уже есть активная задача на этот элемент — не дублируем
        existing = tq.find_active_job_for_element(element_id, project_code)
        if existing:
            return {"status": existing["status"], "job_id": existing["job_id"], "already_queued": True}
 
        # 3. Собираем входные файлы по портам (settings/X_train/Y_train/X_val/Y_val)
        connections = project.get("connections", [])
        conn_by_port = {c["toPort"]: c for c in connections if c["toElement"] == element_id}
 
        input_paths = {}
        input_hashes = {}
        settings_props = {}
 
        for port, role in TEMPLATE_PORT_MAP.items():
            conn = conn_by_port.get(port)
            if not conn:
                if role != 'settings':  # settings опционален, остальные обязательны
                    raise HTTPException(status_code=400,
                                         detail=f"Не подключён вход '{role}' элемента Шаблон")
                continue
            src_id = conn["fromElement"]
            src_elem = elements[src_id]
            from_port = conn.get("fromPort", "out-0")
 
            if role == 'settings':
                # элемент "Настройка" отдаёт свои props напрямую, без файла на диске
                settings_props = dict(src_elem.get("props", {}))
                continue
 
            src_status = get_element_status(src_id, project, config, project_code)
            if not src_status["up_to_date"]:
                raise HTTPException(
                    status_code=409,
                    detail=f"Данные для входа '{role}' не актуальны — примените элемент '{src_id}'"
                )
            path = get_output_file(src_elem, from_port, config, project_code)
            input_paths[role] = path
            input_hashes[src_id] = src_status["hash"]
 
        # 4. Проверяем актуальность обучения — если уже обучено на этих же
        #    данных/дизайне/настройках, не гоняем GPU впустую
        design_project = nn_template.load_design_project(config, design_code)
        d_hash = nn_template.design_hash(design_project)
        train_hash = nn_template.compute_train_hash(design_code, d_hash, settings_props, input_hashes)
 
        meta_path = nn_template.get_model_meta_path(config, project_code, element_id)
        if os.path.exists(meta_path):
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            if meta.get("train_hash") == train_hash and \
               os.path.exists(nn_template.get_model_path(config, project_code, element_id)):
                return {"status": "already_trained", "job_id": None}
 
        # 5. Ставим в очередь
        job_id = tq.enqueue_job(
            config=config,
            project_code=project_code,
            element_id=element_id,
            design_code=design_code,
            user=user,
            settings=settings_props,
            inputs=input_paths,
            train_hash=train_hash,
        )
        # input_hashes понадобятся воркеру для записи в meta.json — допишем в job
        job = tq.get_job(job_id)
        job["input_hashes"] = input_hashes
        tq._write_job(job_id, job)
 
        return {"status": "queued", "job_id": job_id, "queue_position": tq.queue_position(job_id)}
 
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
 
 
# ---------------------------------------------------------------------------
# Статус конкретного job'а (для поллинга прогресса/метрик с фронта)
# ---------------------------------------------------------------------------
 
@app.get("/api/nn/train/status/{job_id}")
async def get_train_job_status(job_id: str):
    job = tq.get_job_with_metrics(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Задача обучения не найдена")
    job["queue_position"] = tq.queue_position(job_id)
    return job
 
 
@app.post("/api/nn/train/cancel/{job_id}")
async def cancel_train_job(job_id: str):
    ok = tq.cancel_job(job_id)
    if not ok:
        raise HTTPException(status_code=409, detail="Задача уже выполняется или завершена — отменить нельзя")
    return {"status": "cancelled"}
 
 
@app.get("/api/nn/train/queue")
async def get_train_queue(config: str = Query(...)):
    jobs = tq.list_jobs(config=config, statuses=["queued", "running"])
    return {"jobs": jobs, "running_job_id": tq.current_running_job_id()}


@app.get("/api/training-params")
async def get_training_params():
    json_path = os.path.join(BASE_DIR, "training_params.json")
    if not os.path.exists(json_path):
        raise HTTPException(status_code=404, detail="training_params.json not found")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data

    






# -----------------------------------------------------------------------------
# Статические файлы (фронтенд)
# -----------------------------------------------------------------------------

WEB_DIR = os.path.normpath(os.path.join(BASE_DIR, "..", "web"))
app.mount("/", StaticFiles(directory=WEB_DIR, html=True), name="web")
print(f"[DEBUG] WEB_DIR: {WEB_DIR}")
for f in os.listdir(WEB_DIR):
    print(f"  {repr(f)}  | exists: {os.path.exists(os.path.join(WEB_DIR, f))}")




